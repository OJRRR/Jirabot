"""批量操作工具 — 批量创建、删除、更新、Excel 导入"""
import json
import os
import logging
from langchain.tools import tool
from config import Config
from utils import extract_issue_key, validate_project_key
from ._lazy import LazyJira
from .task_common import (
    build_issue_fields,
    _resolve_fields_via_meta,
)

_logger = logging.getLogger("jira_bot.task_batch")
jira = LazyJira()


@tool
def batch_create_issues(issues_json: str = "") -> str:
    """批量创建多个 Jira 任务（使用 Jira bulk API，每次最多 50 条并发提交）。"""
    try:
        if not issues_json:
            return json.dumps(
                {"success": False,
                 "error": "请提供 issues_json 参数（JSON 数组字符串）"},
                ensure_ascii=False)

        issues_data = json.loads(issues_json)
        if not isinstance(issues_data, list):
            return json.dumps(
                {"success": False,
                 "error": "issues_json 必须是 JSON 数组"},
                ensure_ascii=False)

        results = []
        errors = []
        batch_size = 50

        for i in range(0, len(issues_data), batch_size):
            batch = issues_data[i:i + batch_size]
            fields_list = []
            batch_meta = []

            for idx, item in enumerate(batch):
                project_key = validate_project_key(
                    item.get("project_key", ""))
                summary = item.get("summary", "")
                issue_type = item.get("issue_type", "Task")
                description = item.get("description", "")
                parent_key = item.get("parent_key")
                epic_link_key = item.get("epic_link_key")
                additional_fields = item.get("additional_fields")

                if not project_key or not summary:
                    errors.append(
                        {"index": i + idx,
                         "error": "缺少 project_key 或 summary"})
                    continue

                if additional_fields:
                    additional_fields = _resolve_fields_via_meta(
                        additional_fields, project_key)

                fields, err = build_issue_fields(
                    project_key=project_key, summary=summary,
                    issue_type=issue_type, description=description,
                    parent_key=parent_key,
                    epic_link_key=epic_link_key,
                    additional_fields=additional_fields)
                if err:
                    errors.append({"index": i + idx, "error": err})
                    continue

                fields_list.append(fields)
                batch_meta.append({"index": i + idx, "summary": summary})

            if fields_list:
                try:
                    created = jira.bulk_create_issues(fields_list)
                    if isinstance(created, list):
                        for j, issue_result in enumerate(created):
                            if j < len(batch_meta):
                                meta = batch_meta[j]
                                if isinstance(issue_result, dict):
                                    key = issue_result.get("key", "未知")
                                    results.append(
                                        {"index": meta["index"],
                                         "key": key,
                                         "summary": meta["summary"],
                                         "status": "created"})
                                else:
                                    errors.append(
                                        {"index": meta["index"],
                                         "error": str(issue_result),
                                         "summary": meta["summary"]})
                    else:
                        for meta in batch_meta:
                            errors.append(
                                {"index": meta["index"],
                                 "error": f"批量创建返回异常: {created}",
                                 "summary": meta["summary"]})
                except Exception as be:
                    for meta in batch_meta:
                        errors.append(
                            {"index": meta["index"],
                             "error": str(be),
                             "summary": meta["summary"]})

        return json.dumps(
            {"success": len(errors) == 0,
             "total": len(issues_data),
             "created": len(results),
             "errors_count": len(errors),
             "results": results, "errors": errors},
            ensure_ascii=False, indent=2)
    except json.JSONDecodeError as e:
        return json.dumps(
            {"success": False,
             "error": f"JSON 解析失败: {str(e)}"},
            ensure_ascii=False)
    except Exception as e:
        return json.dumps({"success": False, "error": str(e)},
                          ensure_ascii=False)


@tool
def batch_delete_issues(issues_json: str = "",
                        delete_subtasks: bool = True,
                        confirmed: bool = False) -> str:
    """批量删除多个 Jira Issue。
    警告：此操作不可逆，请谨慎使用！
    """
    try:
        if not issues_json:
            return json.dumps(
                {"success": False, "error": "请提供 issues_json 参数"},
                ensure_ascii=False)

        issues = json.loads(issues_json)
        if not isinstance(issues, list):
            return json.dumps(
                {"success": False, "error": "issues_json 必须是 JSON 数组"},
                ensure_ascii=False)

        if not confirmed:
            return json.dumps(
                {"success": False,
                 "error": f"⚠️ 即将删除 {len(issues)} 个 Issue，"
                          f"请设置 confirmed=true 确认"},
                ensure_ascii=False)

        results = []
        errors = []
        for item in issues:
            key = extract_issue_key(
                item if isinstance(item, str) else item.get("key", ""))
            if not key:
                errors.append({"item": item, "error": "无效的 issue_key"})
                continue
            try:
                jira.delete_issue(key, delete_subtasks=delete_subtasks)
                results.append({"key": key, "status": "deleted"})
            except Exception as de:
                errors.append({"key": key, "error": str(de)})

        return json.dumps(
            {"success": len(errors) == 0,
             "total": len(issues),
             "deleted": len(results),
             "errors_count": len(errors),
             "results": results, "errors": errors},
            ensure_ascii=False, indent=2)
    except json.JSONDecodeError as e:
        return json.dumps(
            {"success": False, "error": f"JSON 解析失败: {str(e)}"},
            ensure_ascii=False)
    except Exception as e:
        return json.dumps({"success": False, "error": str(e)},
                          ensure_ascii=False)


@tool
def import_from_excel(file_path: str = "", project_key: str = "",
                      summary_col: str = "", issue_type_col: str = "",
                      start_date_col: str = "", end_date_col: str = "",
                      sheet_name: str = "Sheet1", header_row: int = 0,
                      dry_run: bool = False) -> str:
    """从 Excel (.xlsx) 文件导入并批量创建 Jira 任务
    （使用 Jira bulk API，≤50 条/批）。
    自动解析列内容，建议提供列名映射以准确识别。
    """
    try:
        if not file_path or not project_key:
            return json.dumps(
                {"success": False,
                 "error": "请提供 file_path 和 project_key 参数"},
                ensure_ascii=False)

        project_key = validate_project_key(project_key)

        # 支持 uploads/ 目录下的文件路径
        if not os.path.isabs(file_path):
            uploads_dir = Config.UPLOAD_DIR
            full_path = os.path.join(uploads_dir, file_path)
            if os.path.exists(full_path):
                file_path = full_path

        if not os.path.exists(file_path):
            return json.dumps(
                {"success": False, "error": f"文件不存在: {file_path}"},
                ensure_ascii=False)

        import openpyxl
        wb = openpyxl.load_workbook(file_path, data_only=True)
        if sheet_name not in wb.sheetnames:
            return json.dumps(
                {"success": False,
                 "error": f"工作表 '{sheet_name}' 不存在，"
                          f"可用: {wb.sheetnames}"},
                ensure_ascii=False)

        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) <= header_row:
            return json.dumps(
                {"success": False, "error": "文件中没有数据行"},
                ensure_ascii=False)

        headers = [str(h).strip() if h else f"Col{i}"
                   for i, h in enumerate(rows[header_row])]
        data_rows = rows[header_row + 1:]

        # 列名映射
        col_map = {}
        for i, h in enumerate(headers):
            hl = h.lower().replace(" ", "").replace("_", "").replace("-", "")
            col_map[hl] = i

        def _find_col(user_hint: str, aliases: list) -> int:
            if user_hint:
                hl = user_hint.lower().replace(" ", "").replace("_", "")
                if hl in col_map:
                    return col_map[hl]
            for alias in aliases:
                al = alias.lower().replace(" ", "").replace("_", "")
                if al in col_map:
                    return col_map[al]
            return None

        summary_idx = _find_col(summary_col,
                                ["summary", "title", "name", "task", "主题",
                                 "标题", "任务名", "名称"])
        issue_type_idx = _find_col(issue_type_col,
                                   ["issuetype", "type", "问题类型", "类型"])
        start_idx = _find_col(start_date_col,
                              ["targetstart", "startdate", "start",
                               "开始日期", "开始"])
        end_idx = _find_col(end_date_col,
                            ["targetend", "enddate", "due", "end",
                             "结束日期", "截止", "结束"])

        if summary_idx is None:
            return json.dumps(
                {"success": False,
                 "error": f"找不到标题列，可用列: {headers}"},
                ensure_ascii=False)

        issues_to_create = []
        for row_idx, row in enumerate(data_rows):
            summary = str(
                row[summary_idx]).strip() if summary_idx < len(row) and row[
                summary_idx] else ""
            if not summary or summary.lower() in ("none", "null", ""):
                continue

            issue_type = "Task"
            if issue_type_idx is not None and issue_type_idx < len(row) and \
                    row[issue_type_idx]:
                issue_type = str(row[issue_type_idx]).strip() or "Task"

            additional = {}
            if start_idx is not None and start_idx < len(row) and row[
                start_idx]:
                additional["Target Start"] = str(row[start_idx]).strip()
            if end_idx is not None and end_idx < len(row) and row[end_idx]:
                additional["Target End"] = str(row[end_idx]).strip()

            issues_to_create.append({
                "project_key": project_key,
                "summary": summary,
                "issue_type": issue_type,
                "additional_fields": additional if additional else None,
            })

        if dry_run:
            return json.dumps(
                {"success": True, "dry_run": True,
                 "total": len(issues_to_create),
                 "preview": issues_to_create[:10],
                 "message": f"预览模式：将创建 {len(issues_to_create)} 个任务"},
                ensure_ascii=False, indent=2)

        # 批量创建
        results = []
        errors = []
        batch_size = 50
        for i in range(0, len(issues_to_create), batch_size):
            batch = issues_to_create[i:i + batch_size]
            fields_list = []
            batch_meta = []

            for idx, item in enumerate(batch):
                additional_fields = item.get("additional_fields")
                if additional_fields:
                    additional_fields = _resolve_fields_via_meta(
                        additional_fields, project_key)

                fields, err = build_issue_fields(
                    project_key=project_key,
                    summary=item["summary"],
                    issue_type=item["issue_type"],
                    additional_fields=additional_fields)
                if err:
                    errors.append(
                        {"index": i + idx + 1, "error": err,
                         "summary": item["summary"]})
                    continue
                fields_list.append(fields)
                batch_meta.append(
                    {"index": i + idx + 1, "summary": item["summary"]})

            if fields_list:
                try:
                    created = jira.bulk_create_issues(fields_list)
                    if isinstance(created, list):
                        for j, issue_result in enumerate(created):
                            if j < len(batch_meta):
                                meta = batch_meta[j]
                                if isinstance(issue_result, dict):
                                    results.append(
                                        {"index": meta["index"],
                                         "key": issue_result.get("key", "未知"),
                                         "summary": meta["summary"]})
                                else:
                                    errors.append(
                                        {"index": meta["index"],
                                         "error": str(issue_result),
                                         "summary": meta["summary"]})
                except Exception as be:
                    for meta in batch_meta:
                        errors.append(
                            {"index": meta["index"], "error": str(be),
                             "summary": meta["summary"]})

        return json.dumps(
            {"success": len(errors) == 0,
             "total": len(issues_to_create),
             "created": len(results),
             "errors_count": len(errors),
             "results": results, "errors": errors},
            ensure_ascii=False, indent=2)
    except Exception as e:
        return json.dumps({"success": False, "error": str(e)},
                          ensure_ascii=False)


@tool
def batch_update_dates(updates_json: str = "") -> str:
    """批量更新多个 Jira 任务的开始日期和结束日期
    （用于维护 Jira Plan 时间线）。"""
    try:
        if not updates_json:
            return json.dumps(
                {"success": False, "error": "请提供 updates_json 参数"},
                ensure_ascii=False)

        updates = json.loads(updates_json)
        if not isinstance(updates, list):
            return json.dumps(
                {"success": False, "error": "updates_json 必须是 JSON 数组"},
                ensure_ascii=False)

        results = []
        errors = []
        for item in updates:
            key = extract_issue_key(item.get("issue_key", ""))
            if not key:
                errors.append({"item": item, "error": "无效的 issue_key"})
                continue

            fields_to_update = {}
            project_key = item.get("project_key", "")
            if item.get("target_start"):
                fields_to_update["Target Start"] = item["target_start"]
            if item.get("target_end"):
                fields_to_update["Target End"] = item["target_end"]

            if not fields_to_update:
                errors.append({"key": key, "error": "没有要更新的日期字段"})
                continue

            try:
                if project_key:
                    fields_to_update = _resolve_fields_via_meta(
                        fields_to_update, project_key)
                jira.update_issue_field(key, fields_to_update)
                results.append(
                    {"key": key, "updated": list(fields_to_update.keys())})
            except Exception as ue:
                errors.append({"key": key, "error": str(ue)})

        return json.dumps(
            {"success": len(errors) == 0,
             "total": len(updates),
             "updated": len(results),
             "errors_count": len(errors),
             "results": results, "errors": errors},
            ensure_ascii=False, indent=2)
    except json.JSONDecodeError as e:
        return json.dumps(
            {"success": False, "error": f"JSON 解析失败: {str(e)}"},
            ensure_ascii=False)
    except Exception as e:
        return json.dumps({"success": False, "error": str(e)},
                          ensure_ascii=False)


@tool
def batch_update_issues(updates_json: str = "") -> str:
    """批量更新多个 Jira 任务的负责人、状态、优先级。
    一次调用可同时改多个字段，逐条执行，单条失败不影响其他。
    """
    try:
        if not updates_json:
            return json.dumps(
                {"success": False, "error": "请提供 updates_json 参数"},
                ensure_ascii=False)

        updates = json.loads(updates_json)
        if not isinstance(updates, list):
            return json.dumps(
                {"success": False, "error": "updates_json 必须是 JSON 数组"},
                ensure_ascii=False)

        results = []
        errors = []
        for item in updates:
            key = extract_issue_key(item.get("issue_key", ""))
            if not key:
                errors.append({"item": item, "error": "无效的 issue_key"})
                continue

            try:
                # 分配负责人
                if item.get("assignee"):
                    jira.assign_issue(key, item["assignee"])

                # 流转状态
                if item.get("transition"):
                    jira.transition_issue(key, item["transition"])

                # 更新优先级
                if item.get("priority"):
                    jira.update_issue_field(
                        key, {"priority": {"name": item["priority"]}})

                # 更新其他字段
                additional = item.get("additional_fields")
                if additional:
                    project_key = item.get("project_key", "")
                    if project_key:
                        additional = _resolve_fields_via_meta(
                            additional, project_key)
                    jira.update_issue_field(key, additional)

                results.append({"key": key, "status": "updated"})
            except Exception as ue:
                errors.append({"key": key, "error": str(ue)})

        return json.dumps(
            {"success": len(errors) == 0,
             "total": len(updates),
             "updated": len(results),
             "errors_count": len(errors),
             "results": results, "errors": errors},
            ensure_ascii=False, indent=2)
    except json.JSONDecodeError as e:
        return json.dumps(
            {"success": False, "error": f"JSON 解析失败: {str(e)}"},
            ensure_ascii=False)
    except Exception as e:
        return json.dumps({"success": False, "error": str(e)},
                          ensure_ascii=False)
